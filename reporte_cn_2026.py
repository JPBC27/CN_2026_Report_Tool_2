import pandas as pd
import numpy as np
import streamlit as st
from datetime import datetime
import io
import unicodedata
import json
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
try:
    from openpyxl.pivot.table import PivotTable, Reference, PivotField, RowField, ColumnField, DataField
    PIVOT_SUPPORTED = True
except ImportError:
    PIVOT_SUPPORTED = False
from openpyxl.worksheet.table import Table, TableStyleInfo

def normalizar_texto(texto):
    """Elimina acentos y convierte a mayúsculas para comparaciones robustas."""
    if not texto: return ""
    texto = str(texto)
    texto = unicodedata.normalize('NFD', texto).encode('ascii', 'ignore').decode("utf-8")
    return texto.upper().strip()

def limpiar_id(val):
    """Convierte cualquier ID (leído como número, float o texto) a un string de entero puro."""
    if pd.isnull(val) or str(val).strip() == "": return None
    try:
        # Forzar conversión: de cualquier cosa -> float -> int -> string
        # Esto quita el .0, los ceros a la izquierda y espacios.
        return str(int(float(str(val).strip())))
    except (ValueError, TypeError):
        # Si no es numérico, devolver el string limpio
        return str(val).strip()

# --- CONFIGURACIÓN DE CURSOS ---
CONFIG_PATH = "config_cursos.json"

def cargar_config_cursos():
    """Carga la configuración de cursos desde un archivo JSON."""
    if not os.path.exists(CONFIG_PATH):
        # Valores por defecto si no existe el archivo
        config = {
            "ucenco": [
                "PREVENCION DEL HOSTIGAMIENTO SEXUAL LABORAL 2026",
                "PREVENCION DE ENFERMEDADES 2026",
                "REPORTE DE ACCIDENTES E INCIDENTES - 2026"
            ],
            "transversales_sso": [
                "EVALUACION DE SALIDA - HOSTIGAMIENTO SEXUAL 2026",
                "EVALUACION DE SALIDA - PREVENCION EN ENFERMEDADES INFECTOCONTAGIOSAS Y COVID19 2026",
                "EVALUACION DE SALIDA - REPORTE DE ACCIDENTES E INCIDENTES DE TRABAJO 2026",
                "CONOCIENDO E IDENTIFICANDO NUESTROS PELIGROS, RIESGOS, CUMPLIMIENTO DE CONTROLES Y MAPA DE RIESGO 2026",
                "EVALUACION DE SALIDA ESTANDAR DE SEGURIDAD Y SALUD EN ALMACEN Y ABARROTES 2026",
                "EVALUACION DE SALIDA - ESTANDAR DE SEGURIDAD Y SALUD EN EL AREA DE ADUANAS 2026",
                "EVALUACION DE SALIDA ESTANDAR DE SEGURIDAD Y SALUD EN EL AREA DE BAZAR Y TEXTIL 2026",
                "EVALUACION DE SALIDA - ESTANDAR DE SEGURIDAD Y SALUD EN EL AREA DE CAJAS 2026",
                "EVALUACION DE SALIDA - ESTANDAR DE SEGURIDAD Y SALUD EN EL AREA DE ELECTRODOMESTICOS 2025",
                "EVALUACION DE SALIDA - ESTANDAR DE SEGURIDAD Y SALUD EN EL AREA DE RECEPCION DE MERCADERIA 2026",
                "EVALUACION DE SALIDA ESTANDAR DE SEGURIDAD Y SALUD EN PERECIBLES 2026",
                "EVALUACION DE SALIDA - ESTANDAR DE SEGURIDAD Y SALUD EN REPRESENTANTES DE VENTAS 2026",
                # "EVALUACIÓN DE SALIDA - ESPECIALISTA EN FRUTAS Y VERDURAS 2026",
                # "EVALUACIÓN DE SALIDA - ESPECIALISTA EN CALIDAD DE LÁCTEOS Y EMBUTIDOS 2026"
            ]
        }
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return config
    
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def guardar_config_cursos(config):
    """Guarda la configuración de cursos en un archivo JSON."""
    with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

def formatear_nombre_censo(val):
    """Elimina la coma del Censo y asegura un espacio entre apellido y nombre."""
    if pd.isnull(val) or str(val).strip() == "": return ""
    # "PÉREZ, JUAN" -> "PÉREZ JUAN"
    partes = str(val).split(',')
    return " ".join(p.strip() for p in partes if p.strip())

def convertir_fecha_robusta(series, dayfirst=True):
    """
    Convierte una serie a datetime reconociendo formatos latinos, 
    ingleses y números seriales de Excel.
    """
    if series is None: return None
    
    # Intentar conversión estándar de Pandas
    dates = pd.to_datetime(series, dayfirst=dayfirst, errors='coerce')
    
    # Reparar fechas que vienen como números seriales de Excel (ej: 45825.0)
    mask_nat = dates.isna() & series.notna() & (series.astype(str).str.strip() != "")
    if mask_nat.any():
        def convert_excel_num(val):
            try:
                f = float(val)
                # Excel serial dates are usually > 30000 (1982+)
                if 30000 < f < 60000:
                    return pd.to_datetime(f, unit='D', origin='1899-12-30')
            except:
                pass
            return pd.NaT
        
        repaired = series[mask_nat].apply(convert_excel_num)
        dates.loc[mask_nat] = repaired
        
    return dates

def cargar_excel_inteligente(file, nombres_preferidos=None):
    """
    Carga un Excel buscando automáticamente la hoja y fila del encabezado.
    Prioriza nombres de hojas si se especifican.
    """
    try:
        xls = pd.ExcelFile(file)
        sheets = xls.sheet_names
        
        search_order = []
        if nombres_preferidos:
            for pref in nombres_preferidos:
                for s in sheets:
                    if pref.upper() in s.upper() and s not in search_order:
                        search_order.append(s)
        for s in sheets:
            if s not in search_order: search_order.append(s)
        
        seeds = [
            'SAP', 'PERS', 'NOMBRE', 'ALTA', 'INGRES', 'CD', 'UBIC', 
            'DNI', 'DOC', 'ID', 'FUNCION'
        ]

        for sheet_name in search_order:
            df_raw = pd.read_excel(xls, header=None, sheet_name=sheet_name, dtype=str, nrows=50)
            for i in range(len(df_raw)):
                row_values = [normalizar_texto(str(val)) for val in df_raw.iloc[i].values if val is not None]
                matches = sum(1 for s in seeds if any(s in val for val in row_values))
                if matches >= 3:
                    st.write(f"📝 Cabecera detectada en fila {i} (hoja: {sheet_name})")
                    return pd.read_excel(xls, header=i, sheet_name=sheet_name, dtype=str)
            
            # Fallback: primera fila con datos reales
            for i in range(len(df_raw)):
                if df_raw.iloc[i].dropna().count() >= 5:
                    st.write(f"📝 Cabecera detectada en fila {i} (hoja: {sheet_name}) por densidad de datos.")
                    return pd.read_excel(xls, header=i, sheet_name=sheet_name, dtype=str)
        return pd.read_excel(xls, sheet_name=0, dtype=str)
    except Exception as e:
        st.error(f"Error cargando el archivo {getattr(file, 'name', 'Excel')}: {e}")
        return pd.DataFrame()

def procesar_archivo_dotacion(file_dotacion):
    """Procesa un único archivo de Dotación General usando cabeceras fijas."""
    st.write(f"📥 Procesando Dotación (Cabeceras Fijas): `{file_dotacion.name}`")
    df_dot = cargar_excel_inteligente(file_dotacion, nombres_preferidos=['DOTACION', 'GLOBAL'])
    
    if df_dot.empty: return None, None, {}

    n_ini = len(df_dot)
    st.write(f"📊 Dotación: Se leyeron **{n_ini}** registros iniciales.")

    # Validar existencia de Alta
    if 'Alta' not in df_dot.columns:
        st.error(f"⚠️ No se encontró la columna 'Alta' en `{file_dotacion.name}`.")
        return None, None, {}

    # Conversión y Limpieza
    df_dot['Alta_Date'] = convertir_fecha_robusta(df_dot['Alta'], dayfirst=True)
    if 'Nº pers.' in df_dot.columns:
        df_dot['Nº pers.'] = df_dot['Nº pers.'].apply(limpiar_id)
        df_dot = df_dot.drop_duplicates(subset=['Nº pers.'], keep='first')

    # Filtros de Negocio (Usando Clasificación y Unidad de Negocio)
    def filter_dotacion(row):
        clasif = str(row.get('Clasificación', '')).upper()
        unidad = str(row.get('Unidad de Negocio', '')).upper()
        if 'PERECIBLES' in unidad and clasif != 'FABRICA': return False
        if 'ECOMMERCE' in unidad and clasif != 'LOCALES': return False
        return True
    
    df_dot = df_dot[df_dot.apply(filter_dotacion, axis=1)]
    
    n_fin = len(df_dot)
    if n_ini != n_fin:
        st.write(f"📉 Filtros de Negocio (Perecibles/eCommerce): Quedan **{n_fin}** registros (Se filtraron {n_ini - n_fin}).")
    
    # Metadatos para Censo
    mask_dates = (df_dot['Alta_Date'].dt.year < 2100) & (df_dot['Alta_Date'].dt.year > 1900)
    max_alta = df_dot[mask_dates]['Alta_Date'].max()
    
    df_dot['Alta'] = df_dot['Alta_Date'].dt.strftime('%d/%m/%Y').fillna(df_dot['Alta'])
    
    subdir_map = {}
    if 'SubPer' in df_dot.columns and 'Subdivisión de personal' in df_dot.columns:
        subdir_map = df_dot.dropna(subset=['Subdivisión de personal']).groupby('SubPer')['Subdivisión de personal'].first().to_dict()
    
    return df_dot, max_alta, subdir_map
    """Procesa un único archivo de Dotación General y retorna el DataFrame y sus metadatos."""
    st.write(f"📥 Procesando Dotación: `{file_dotacion.name}`")
    df_dotation = cargar_excel_inteligente(file_dotacion, nombres_preferidos=['DOTACION', 'GLOBAL'])
    
    col_alta = obtener_columna(df_dotation, 'Alta')
    if not col_alta:
        st.error(f"⚠️ No se encontró la columna de 'Alta' en `{file_dotacion.name}`.")
        return None, None, {}

    # Conversión y Limpieza
    df_dotation['Alta_Date'] = convertir_fecha_robusta(df_dotation[col_alta])
    col_sap_dot = obtener_columna(df_dotation, 'Nº pers.')
    if col_sap_dot:
        df_dotation[col_sap_dot] = df_dotation[col_sap_dot].apply(limpiar_id)
        df_dotation = df_dotation.drop_duplicates(subset=[col_sap_dot], keep='first')

    # Filtros de Negocio
    col_clasif = obtener_columna(df_dotation, 'Clasificacion')
    col_unidad = obtener_columna(df_dotation, 'Unidad de Negocio')
    def filter_dotacion(row):
        clasif = str(row.get(col_clasif, '')).upper() if col_clasif else ""
        unidad = str(row.get(col_unidad, '')).upper() if col_unidad else ""
        if 'PERECIBLES' in unidad and clasif != 'FABRICA': return False
        if 'ECOMMERCE' in unidad and clasif != 'LOCALES': return False
        return True
    
    df_dotation = df_dotation[df_dotation.apply(filter_dotacion, axis=1)]
    
    # Metadatos para Censo
    mask_dates = (df_dotation['Alta_Date'].dt.year < 2100) & (df_dotation['Alta_Date'].dt.year > 1900)
    max_alta = df_dotation[mask_dates]['Alta_Date'].max()
    
    df_dotation[col_alta] = df_dotation['Alta_Date'].dt.strftime('%d/%m/%Y').fillna(df_dotation[col_alta])
    
    col_sp_dot = obtener_columna(df_dotation, 'SubPer')
    col_sub_dot = obtener_columna(df_dotation, 'Subdivisión de personal')
    subdir_map = {}
    if col_sp_dot and col_sub_dot:
        subdir_map = df_dotation.dropna(subset=[col_sub_dot]).groupby(col_sp_dot)[col_sub_dot].first().to_dict()
    
    # Renombrado Estándar
    mapping = {
        col_sap_dot: 'Nº pers.',
        obtener_columna(df_dotation, 'Nombre del empleado o candidat'): 'Nombre del empleado o candidat',
        col_alta: 'Alta',
        col_sp_dot: 'SubPer',
        col_sub_dot: 'Subdivisión de personal',
        obtener_columna(df_dotation, 'Función'): 'Función',
        obtener_columna(df_dotation, 'ID Number'): 'ID Number'
    }
    df_dotation.rename(columns={k: v for k, v in mapping.items() if k}, inplace=True)
    
    return df_dotation, max_alta, subdir_map

def procesar_archivo_censo(file_censo, max_alta, subdir_map):
    """Procesa un único archivo de Censo mapeando desde cabeceras fijas."""
    st.write(f"🔍 Procesando Censo (Cabeceras Fijas): `{file_censo.name}`")
    df_c = cargar_excel_inteligente(file_censo, nombres_preferidos=['CENSO'])
    if df_c.empty: return pd.DataFrame()

    if 'Fecha Ingreso Planilla' not in df_c.columns:
        st.warning(f"❌ Sin columna 'Fecha Ingreso Planilla' en `{file_censo.name}`.")
        return pd.DataFrame()

    # Conversión de Fecha (Forzando mm/dd/yyyy para Censo si es texto)
    df_c['Fecha_Dt'] = convertir_fecha_robusta(df_c['Fecha Ingreso Planilla'], dayfirst=False)
    if not pd.isnull(max_alta):
        df_c = df_c[df_c['Fecha_Dt'] > max_alta]
    
    if 'Cd Ubicación' in df_c.columns:
        df_c['Cd Ubicación'] = df_c['Cd Ubicación'].fillna('').astype(str).str.strip().str.upper()
        df_c = df_c[~df_c['Cd Ubicación'].str.startswith(('B', 'C'))]
        df_c = df_c[~df_c['Cd Ubicación'].isin(['H099', 'HPCM'])]
    
    n_c_pre_area = len(df_c)
    
    # Filtro de Área de Trabajo Flexibilizado (Logística, Operaciones, Centro de Distribución)
    areas_validas = ["LOGISTICA", "OPERACIONES", "CENTRO DE DISTRIBUCION"]
    if 'Área de Trabajo' in df_c.columns:
        def matches_area(val):
            if pd.isnull(val): return False
            n = normalizar_texto(val)
            return any(a in n for a in areas_validas)
        df_c = df_c[df_c['Área de Trabajo'].apply(matches_area)]
        
    n_c_post_area = len(df_c)
    if n_c_pre_area != n_c_post_area:
        st.write(f"🔎 Censo: Se encontraron **{n_c_post_area}** registros en áreas válidas de {n_c_pre_area} disponibles.")
    
    if df_c.empty: return pd.DataFrame()

    # Mapeo Directo (Censo -> Final)
    df_res = pd.DataFrame({
        'Nº pers.': df_c['Cod. Sap'].apply(limpiar_id) if 'Cod. Sap' in df_c.columns else "",
        'Nombre del empleado o candidat': df_c['Nombre'].apply(formatear_nombre_censo) if 'Nombre' in df_c.columns else "",
        'Alta': df_c['Fecha_Dt'].dt.strftime('%d/%m/%Y') if not df_c['Fecha_Dt'].isnull().all() else "",
        'SubPer': df_c['Cd Ubicación'] if 'Cd Ubicación' in df_c.columns else "",
        'Función': df_c['Puesto'].astype(str).str.strip() if 'Puesto' in df_c.columns else "",
        'ID Number': df_c['Doc ID'].apply(limpiar_id) if 'Doc ID' in df_c.columns else ""
    })
    
    df_res['Subdivisión de personal'] = df_res['SubPer'].map(subdir_map).fillna("")
    return df_res

def procesar_archivo_adicional(file_adicional):
    """Extrae diccionarios de mapeo del archivo de Información Adicional."""
    st.write(f"📊 Extrayendo Info Adicional: `{file_adicional.name}`")
    try:
        xls = pd.ExcelFile(file_adicional)
        info_dict = {}
        sheets = xls.sheet_names
        
        # 1. Hoja Cod Organicos (Banderas)
        sh_org = next((s for s in sheets if 'Cod Organicos' in s), None)
        if sh_org:
            df_org = pd.read_excel(xls, sheet_name=sh_org, header=None, dtype=str)
            info_dict['map_org'] = df_org.set_index(0)[1].to_dict()
            
        # 2. Hoja Secciones (Funciones -> Secciones)
        sh_sec = next((s for s in sheets if s.lower() == 'secciones'), None)
        if sh_sec:
            df_sec = pd.read_excel(xls, sheet_name=sh_sec, dtype=str)
            df_sec.columns = [str(c).strip() for c in df_sec.columns]
            info_dict['df_sec'] = df_sec

        # 3. Hoja Distr Jefes Tiendas (SubPer -> Ciudad, Formato, G Zonal, JGH, JGHZ)
        sh_jef = next((s for s in sheets if 'distr jefes tiendas' in s.lower()), None)
        if sh_jef:
            df_jef = pd.read_excel(xls, sheet_name=sh_jef, dtype=str)
            df_jef.columns = [str(c).strip() for c in df_jef.columns]
            info_dict['df_jefes'] = df_jef
            
        # 4. Hoja Lista de cursos (Mapeo de Capacitaciones)
        sh_cursos = next((s for s in sheets if 'lista de cursos' in s.lower()), None)
        if sh_cursos:
            df_m = pd.read_excel(xls, sheet_name=sh_cursos, dtype=str)
            df_m.columns = [str(c).strip() for c in df_m.columns]
            
            # Nombre de la columna de grupo solicitado
            col_grupo = 'Grupo para resultado excel'
            if col_grupo not in df_m.columns:
                # Fallback a 'Grupo' si no existe la nueva
                col_grupo = 'Grupo' if 'Grupo' in df_m.columns else col_grupo
                if col_grupo not in df_m.columns:
                    df_m[col_grupo] = "SIN GRUPO"

            # --- NUEVO: Extraer listas de cursos dinámicas ---
            if 'Nombre Capa en UCENCO' in df_m.columns:
                list_u = df_m['Nombre Capa en UCENCO'].dropna().unique().tolist()
                info_dict['lista_ucenco'] = [c.strip() for c in list_u if str(c).strip() and str(c).upper() != 'NAN']
            
            if 'Nombre Capa en Campus Cencosud' in df_m.columns:
                list_c = df_m['Nombre Capa en Campus Cencosud'].dropna().unique().tolist()
                info_dict['lista_campus'] = [c.strip() for c in list_c if str(c).strip() and str(c).upper() != 'NAN']

            info_dict['df_matriz_cursos'] = df_m
            info_dict['col_grupo_real'] = col_grupo # Para que otras funciones sepan cuál usar
            
            # --- NUEVO: Extraer colores con OpenPyXL ---
            try:
                # Resetear puntero del buffer para lectura con openpyxl
                if hasattr(file_adicional, 'seek'):
                    file_adicional.seek(0)
                wb_col = load_workbook(file_adicional, data_only=True)
                ws_col = wb_col[sh_cursos]
                headers_col = [str(cell.value).strip() for cell in ws_col[1]]
                
                if col_grupo in headers_col and 'Color' in headers_col:
                    idx_g = headers_col.index(col_grupo)
                    idx_c = headers_col.index('Color')
                    idx_td = headers_col.index('Grupo TD') if 'Grupo TD' in headers_col else -1
                    
                    map_colores = {}
                    map_td = {} # Nuevo: Mapeo de Curso -> Grupo TD
                    for row in ws_col.iter_rows(min_row=2):
                        nombre_rep = str(row[headers_col.index('Nombre para el Reporte')].value).strip() if 'Nombre para el Reporte' in headers_col else None
                        g_val = str(row[idx_g].value).strip().upper()
                        
                        if g_val and g_val != 'NONE' and g_val != 'NAN':
                            # Prioridad 1: Color de fondo (Relleno) de la celda
                            fill = row[idx_c].fill
                            hex_c = None
                            if fill and hasattr(fill, 'start_color'):
                                rgb = fill.start_color.rgb
                                if rgb and isinstance(rgb, str) and rgb != "00000000":
                                    hex_c = rgb[2:] if len(rgb) == 8 else rgb
                            
                            # Prioridad 2: Valor de la celda (si es un código Hex escrito o "Sin Color")
                            if not hex_c or hex_c == "000000":
                                val_c = str(row[idx_c].value).strip().replace("#", "").upper()
                                if len(val_c) == 6 and all(c in "0123456789ABCDEF" for c in val_c):
                                    hex_c = val_c
                            
                            if hex_c and hex_c != "000000":
                                map_colores[g_val] = hex_c
                        
                        # Capturar Grupo TD para el Dashboard
                        if nombre_rep and idx_td != -1:
                            td_val = str(row[idx_td].value).strip()
                            if td_val and td_val != 'None' and td_val != 'nan':
                                map_td[nombre_rep] = td_val
                                
                    info_dict['map_colores_grupos'] = map_colores
                    info_dict['map_grupo_td'] = map_td
            except Exception as e:
                print(f"Error extrayendo colores: {e}")
            
        # 5. Hoja Filtros (Nuevo: Dinámico para NA y Eliminación)
        sh_filtros = next((s for s in sheets if 'filtros' in s.lower()), None)
        if sh_filtros:
            df_filtros = pd.read_excel(xls, sheet_name=sh_filtros, dtype=str)
            df_filtros.columns = [str(c).strip() for c in df_filtros.columns]
            # Normalizar nombres de columnas para evitar problemas de espacios
            # Esperado: Curso, Campo, Operador, Valores (;), Operador Lógico, Tipo
            info_dict['df_filtros'] = df_filtros
            st.write(f"⚙️ Se cargó la hoja de **Filtros** dinámica.")
            
        return info_dict
    except Exception as e:
        st.error(f"Error en Info Adicional: {e}")
        return {}

columnas_finales = [
    'Nº pers.', 'Nombre del empleado o candidat', 'Alta', 'SubPer', 
    'Subdivisión de personal', 'Función', 'ID Number', 'Bandera', 
    'Ciudad', 'Formato', 'Gte. Zonal', 'JGH', 'JGH ZONAL', 'Sección'
]

def enriquecer_final(df_final, info_adicional):
    """Genera las 14 columnas finales exactas solicitadas."""
    # Asegurar que todas las columnas existan
    for c in columnas_finales:
        if c not in df_final.columns: df_final[c] = ""

    # --- FILTRO DE INTEGRIDAD (Limpiar filas de resumen/vacías) ---
    # Solo conservamos filas que tengan un ID Number y un Nombre real.
    # Esto elimina filas tipo "Total", sumatorias o separadores de Excel.
    df_final = df_final[
        df_final['Nombre del empleado o candidat'].astype(str).str.len() > 3
    ].copy()
    
    # Eliminar duplicados residuales y filas con ID nulo
    df_final = df_final.dropna(subset=['ID Number'])
    df_final = df_final[df_final['ID Number'].astype(str).str.strip() != ""].copy()

    # Aplicar Banderas
    if 'map_org' in info_adicional:
        df_final['Bandera'] = df_final['SubPer'].map(info_adicional['map_org']).fillna(df_final['Bandera'])

    # Aplicar Secciones (Función -> Funciones)
    if 'df_sec' in info_adicional:
        df_sec = info_adicional['df_sec']
        if 'Funciones' in df_sec.columns and 'Secciones' in df_sec.columns:
            df_sec['F_N'] = df_sec['Funciones'].apply(normalizar_texto)
            map_sec = df_sec.drop_duplicates('F_N').set_index('F_N')['Secciones'].to_dict()
            df_final['Sección'] = df_final['Función'].apply(normalizar_texto).map(map_sec).fillna(df_final['Sección'])

    # Aplicar Jefaturas (SubPer -> Codigo2)
    if 'df_jefes' in info_adicional:
        df_jef = info_adicional['df_jefes'].copy()
        
        # Normalizar nombres de columnas de la hoja de jefes para búsqueda robusta
        # "G Zonal" podría venir como "G. ZONAL", "G_ZONAL", etc.
        def norm_col(c): return normalizar_texto(c).replace(".", "").replace("_", " ")
        df_jef.columns = [norm_col(c) for c in df_jef.columns]
        
        if 'CODIGO2' in df_jef.columns:
            # Asegurar mapeo de Codigo2 como string limpio para match con SubPer
            df_jef['Cod2_Clean'] = df_jef['CODIGO2'].apply(limpiar_id)
            
            # Mapeos específicos solicitados:
            # {Columna Final: [Nombres Posibles en Excel (Normalizados)]}
            mapeos = {
                'Ciudad': ['CIUDAD', 'CIU'],
                'Formato': ['FORMATO', 'FORM'],
                'Gte. Zonal': ['GTE ZONAL', 'G ZONAL', 'GERENTE ZONAL'],
                'JGH': ['JGH'],
                'JGH ZONAL': ['JGH ZONAL']
            }
            
            for col_final, nombres_posibles in mapeos.items():
                # Encontrar la columna en el Excel que coincida con alguno de los nombres posibles
                col_excel_real = next((c for c in df_jef.columns if c in nombres_posibles), None)
                
                if col_excel_real:
                    # Crear mapa basado en Codigo2 (Cod2_Clean)
                    mapping_dict = df_jef.drop_duplicates('Cod2_Clean').set_index('Cod2_Clean')[col_excel_real].to_dict()
                    # Aplicar al consolidado (Limpiar SubPer para asegurar match con Cod2_Clean)
                    vals_mapped = df_final['SubPer'].apply(limpiar_id).map(mapping_dict)
                    df_final[col_final] = vals_mapped.fillna(df_final[col_final])

    # Asegurar formato entero para IDs en el reporte final
    for col_id in ['Nº pers.', 'ID Number']:
        if col_id in df_final.columns:
            df_final[col_id] = pd.to_numeric(df_final[col_id], errors='coerce').astype('Int64')

    df_final = df_final.drop_duplicates(subset=['Nº pers.'], keep='first')

    # --- NUEVOS FILTROS DE LIMPIEZA FINAL ---
    n_pre_filt = len(df_final)
    
    # 1. Filtro por Código (SubPer): HPMC, HPCM, H099 o inicia con B o C
    if 'SubPer' in df_final.columns:
        df_final['SubPer_Clean'] = df_final['SubPer'].astype(str).str.strip().str.upper()
        mask_codigos = (
            df_final['SubPer_Clean'].isin(['HPMC', 'HPCM', 'H099']) |
            df_final['SubPer_Clean'].str.startswith(('B', 'C'))
        )
        df_final = df_final[~mask_codigos].copy()
        df_final.drop(columns=['SubPer_Clean'], inplace=True)
        
    n_post_filt = len(df_final)
    if n_pre_filt != n_post_filt:
        st.write(f"✂️ Limpieza Final: Se eliminaron **{n_pre_filt - n_post_filt}** registros (Filtros de Códigos).")

    return df_final[columnas_finales]

def procesar_archivo_cesados(file_cesados):
    """Procesa el archivo Maestro Cesados buscando Doc ID y Fecha Baja."""
    st.write(f"📉 Procesando Maestro Cesados: `{file_cesados.name}`")
    df_ces = cargar_excel_inteligente(file_cesados, nombres_preferidos=['CESADOS'])
    if df_ces.empty: return pd.DataFrame()

    # Campos requeridos: Fecha Baja, Doc ID
    if 'Fecha Baja' not in df_ces.columns or 'Doc ID' not in df_ces.columns:
        st.error(f"⚠️ El archivo `{file_cesados.name}` no tiene las cabeceras requeridas ('Fecha Baja', 'Doc ID').")
        return pd.DataFrame()

    df_ces['Fecha_Baja_Dt'] = convertir_fecha_robusta(df_ces['Fecha Baja'])
    df_ces['Doc ID Clean'] = df_ces['Doc ID'].apply(limpiar_id)
    
    return df_ces[['Doc ID Clean', 'Fecha_Baja_Dt']].dropna(subset=['Doc ID Clean'])

def aplicar_purga_cesados(df_final, df_cesados):
    """
    Elimina registros del consolidado comparando Alta vs Fecha de Baja.
    Conserva si Alta > Fecha Baja (Reingreso).
    """
    if df_cesados.empty or df_final.empty: return df_final

    st.write("📈 Validando cronología de reingresos vs bajas...")
    
    # 1. Tomar la fecha de baja más reciente por ID para evitar errores con historiales
    # Agrupamos por 'Doc ID Clean' que ya es string unificado
    df_cesados_max = df_cesados.groupby('Doc ID Clean')['Fecha_Baja_Dt'].max().reset_index()
    ids_cesados_map = df_cesados_max.set_index('Doc ID Clean')['Fecha_Baja_Dt'].to_dict()
    
    # 2. Preparar ID de comparación en el consolidado (Str format)
    df_final['ID_Temp_Str'] = df_final['ID Number'].apply(limpiar_id)
    
    # 3. Preparar Fecha de Alta para comparación numérica
    df_final['Alta_Dt_Temp'] = pd.to_datetime(df_final['Alta'], dayfirst=True, errors='coerce')
    
    # Debug visual para el usuario
    with st.expander("🔍 Verificación de Formatos (ID y Fechas)"):
        c1, c2 = st.columns(2)
        c1.write("Ejemplos en Consolidado:")
        c1.code(list(df_final['ID_Temp_Str'].dropna().unique()[:5]))
        c2.write("Ejemplos en Maestro Cesados:")
        c2.code(list(ids_cesados_map.keys())[:5])

    def stay_logic(row):
        id_str = row['ID_Temp_Str']
        if id_str not in ids_cesados_map:
            return True # No está en cesados, se queda.
            
        fecha_baja = ids_cesados_map[id_str]
        fecha_alta = row['Alta_Dt_Temp']
        
        # Si no hay fecha de alta o baja válida, por seguridad lo mantenemos
        if pd.isnull(fecha_alta) or pd.isnull(fecha_baja):
            return True
            
        # REGLA: Si Alta es mayor que la Baja, es un reingreso -> Se queda.
        return fecha_alta > fecha_baja

    # 4. Aplicar filtrado e identificar eliminados
    df_merge = df_final.copy()
    df_merge['Stay'] = df_merge.apply(stay_logic, axis=1)
    
    df_final_filtered = df_merge[df_merge['Stay']].copy()
    df_eliminados = df_merge[~df_merge['Stay']].copy()
    
    eliminados_count = len(df_eliminados)
    if eliminados_count > 0:
        st.success(f"✅ Filtro cronológico aplicado: Se eliminaron **{eliminados_count}** cesados reales.")
        
        # Cruzar con df_cesados para recuperar la columna 'Fecha Baja' original si se desea,
        # o usar la fecha máxima encontrada. Aquí usaremos la max_baja del mapa.
        df_eliminados['Fecha de Baja Confirmada'] = df_eliminados['ID_Temp_Str'].map(
            lambda x: ids_cesados_map[x].strftime('%d/%m/%Y') if x in ids_cesados_map else ""
        )
    else:
        st.info("ℹ️ No se detectaron cesados recientes que requieran eliminación.")
    
    # Limpiar columnas temporales
    df_final_filtered = df_final_filtered.drop(columns=['ID_Temp_Str', 'Alta_Dt_Temp', 'Stay'])
    
    # El reporte de eliminados solo debe contener columnas relevantes y el formato de fecha solicitado
    cols_eliminados = ['Nº pers.', 'Nombre del empleado o candidat', 'Alta', 'ID Number', 'Fecha de Baja Confirmada']
    df_reporte_eliminados = df_eliminados[cols_eliminados] if not df_eliminados.empty else pd.DataFrame()
    
    return df_final_filtered, df_reporte_eliminados

# --- MÓDULO DE CAPACITACIONES (PROCESO INTERNO) ---

def procesar_ucenco(file_ucenco, cursos_validos_custom=None):
    """
    Procesa archivo Ucenco:
    - Estado 2026: Basado en 'Fecha del examen' y 'Estado del expediente'.
    """
    st.write(f"🎓 Procesando Ucenco: `{file_ucenco.name}`")
    df = cargar_excel_inteligente(file_ucenco)
    if df.empty: return pd.DataFrame()

    # Filtro de Cursos Solicitados
    if cursos_validos_custom is not None:
        cursos_validos = [normalizar_texto(c) for c in cursos_validos_custom]
    else:
        config = cargar_config_cursos()
        cursos_validos = [normalizar_texto(c) for c in config.get("ucenco", [])]
    
    # Normalizar columnas y Nombres de Cursos
    df.columns = [str(c).strip() for c in df.columns]
    
    def logic_ucenco(row):
        fecha = str(row.get('Fecha del examen', '')).strip()
        estado_exp = normalizar_texto(str(row.get('Estado del expediente', '')))
        
        # SI tiene fecha:
        if fecha != "" and fecha.lower() != "nan":
            if "PROCESO" in estado_exp: return "Desaprobado"
            if "REALIZADO" in estado_exp: return "Terminado"
        
        # SI NO tiene fecha:
        return "Pendiente"

    # Filtrar solo cursos 2026
    df['Curso_Norm'] = df['Título de la capacitación'].apply(normalizar_texto)
    df = df[df['Curso_Norm'].isin(cursos_validos)].copy()
    
    # Aplicar lógica y limpiar ID
    df['Estado_Calculado'] = df.apply(logic_ucenco, axis=1)
    df['ID_Clean'] = df['DNI'].apply(limpiar_id)
    df['Título de la capacitación'] = df['Título de la capacitación'].astype(str).str.strip()
    
    return df[['ID_Clean', 'Título de la capacitación', 'Estado_Calculado']]

def procesar_capacitacion_comun(file, tipo_nombre, cursos_validos_custom=None):
    """
    Procesa Transversales y SSO (Misma lógica):
    - Estado 2026: Basado en 'Fecha de finalización de expediente' y 'Estado del expediente'.
    """
    st.write(f"🎓 Procesando {tipo_nombre}: `{file.name}`")
    df = cargar_excel_inteligente(file)
    if df.empty: return pd.DataFrame()

    df.columns = [str(c).strip() for c in df.columns]

    # Filtros de Cursos Transversales / SSO
    if cursos_validos_custom is not None:
        cursos_all = [normalizar_texto(c) for c in cursos_validos_custom]
    else:
        config = cargar_config_cursos()
        cursos_all = [normalizar_texto(c) for c in config.get("transversales_sso", [])]

    def logic_trans_sso(row):
        f_fin = str(row.get('Fecha de finalización de expediente', '')).strip()
        estado_exp = normalizar_texto(str(row.get('Estado del expediente', '')))
        
        if f_fin != "" and f_fin.lower() != "nan":
            return "Terminado"
        
        if "FALLO" in estado_exp: return "Desaprobado"
        if any(x in estado_exp for x in ["PROGRESO", "REGISTRADO"]): return "Pendiente"
        
        return "Pendiente"

    # Filtrar solo cursos válidos
    df['Curso_Norm'] = df['Título de la capacitación'].apply(normalizar_texto)
    # Remove special chars – punctuation for better match
    df['Curso_Norm'] = df['Curso_Norm'].str.replace('–', '-', regex=False).str.replace('  ', ' ', regex=False)
    df = df[df['Curso_Norm'].isin(cursos_all)].copy()
    
    # Aplicar lógica y limpiar ID
    df['Estado_Calculado'] = df.apply(logic_trans_sso, axis=1)
    df['ID_Clean'] = df['Identificación de usuario'].apply(limpiar_id)
    df['Título de la capacitación'] = df['Título de la capacitación'].astype(str).str.strip()
    
    return df[['ID_Clean', 'Título de la capacitación', 'Estado_Calculado']]

def aplicar_filtros_dinamicos(df, df_filtros, modo="No Aplica", curso=None):
    """
    Aplica filtros dinámicos basados en la hoja 'Filtros'.
    Modo: "Eliminar" (Global) o "No Aplica" (Por Curso).
    """
    if df_filtros is None or df_filtros.empty:
        return df if modo == "Eliminar" else pd.Series(False, index=df.index)

    df_f = df_filtros.copy()
    
    # Filtrar por tipo y, si corresponde, por curso
    if modo == "Eliminar":
        filtros = df_f[df_f['Tipo'].fillna('').str.upper().str.strip() == "ELIMINAR"]
    else:
        filtros = df_f[df_f['Tipo'].fillna('').str.upper().str.strip() == "NO APLICA"]
        if curso:
            curso_norm = normalizar_texto(curso)
            filtros = filtros[filtros['Curso'].apply(normalizar_texto) == curso_norm]
    
    if filtros.empty:
        return df if modo == "Eliminar" else pd.Series(False, index=df.index)

    # Mapeo de operadores (Soporte Español/Inglés)
    op_map = {
        'EN': 'IN', 'IN': 'IN', 'ESTA EN': 'IN', 'ESTA EN (;)': 'IN',
        'NO EN': 'NOT IN', 'NOT IN': 'NOT IN', 'NO ESTA EN': 'NOT IN',
        'IGUAL': 'EQUALS', 'EQUALS': 'EQUALS', '==': 'EQUALS', 'IGUAL A': 'EQUALS',
        'CONTIENE': 'CONTAINS', 'CONTAINS': 'CONTAINS', 'CONTENIDO': 'CONTAINS', 'INCLUYE': 'CONTAINS'
    }

    def evaluar_fila_filtro(df_local, f_row):
        campo_excel = str(f_row['Campo']).strip()
        campo = next((c for c in df_local.columns if normalizar_texto(c) == normalizar_texto(campo_excel)), None)
        
        if not campo:
            return pd.Series(False, index=df_local.index)
        
        op_raw = str(f_row['Operador']).upper().strip()
        op = op_map.get(op_raw, 'IN')
        
        valores_raw = str(f_row['Valores (;)']).split(';')
        valores = [normalizar_texto(v) for v in valores_raw if str(v).strip() != "" and str(v).upper() != 'NAN']
        
        col_vals = df_local[campo].apply(normalizar_texto)
        
        if op == 'IN':
            return col_vals.isin(valores)
        elif op == 'NOT IN':
            return ~col_vals.isin(valores)
        elif op == 'EQUALS':
            val_to_match = valores[0] if valores else ""
            return col_vals == val_to_match
        elif op == 'CONTAINS':
            val_to_match = valores[0] if valores else ""
            return col_vals.str.contains(val_to_match, na=False)
        return pd.Series(False, index=df_local.index)

    mask_final = None
    ultimo_op_logico = "OR"

    for _, f_row in filtros.iterrows():
        mask_actual = evaluar_fila_filtro(df, f_row)
        
        if mask_final is None:
            mask_final = mask_actual
        else:
            if ultimo_op_logico == "AND":
                mask_final = mask_final & mask_actual
            else:
                mask_final = mask_final | mask_actual
        
        ultimo_op_logico = str(f_row.get('Operador Lógico', 'OR')).upper().strip()
        if ultimo_op_logico not in ["AND", "OR"]:
            ultimo_op_logico = "OR"

    if modo == "Eliminar":
        if mask_final is not None:
            return df[~mask_final].copy()
        return df
    else:
        return mask_final if mask_final is not None else pd.Series(False, index=df.index)

def consolidar_capacitaciones(df_final, df_ucenco, df_campus, info_adicional):
    """
    Integra los datos de capacitación en el DataFrame consolidado usando jerarquía de estados.
    """
    if df_final.empty: return df_final
    
    st.write("🔄 Aplicando Jerarquía de Capacitaciones...")

    # 0. Aplicar Filtros Globales de "Eliminar"
    df_filtros = info_adicional.get('df_filtros', pd.DataFrame())
    if not df_filtros.empty:
        n_pre_del = len(df_final)
        df_final = aplicar_filtros_dinamicos(df_final, df_filtros, modo="Eliminar")
        n_post_del = len(df_final)
        if n_pre_del != n_post_del:
            st.warning(f"🚫 Filtros de Eliminación: Se eliminaron **{n_pre_del - n_post_del}** registros del consolidado.")
    
    # 1. Preparar Matriz de Mapeo
    df_m = info_adicional.get('df_matriz_cursos', pd.DataFrame())
    if df_m.empty:
        st.warning("⚠️ No se encontró la hoja 'Lista de cursos' en Información Adicional. Se usará el nombre original.")
        # Creamos una estructura mínima si no existe
        map_u, map_c, cols_reporte = {}, {}, []
    else:
        # Limpiar espacios en los valores de búsqueda de la matriz
        df_m['Nombre Capa en UCENCO'] = df_m['Nombre Capa en UCENCO'].astype(str).str.strip()
        df_m['Nombre Capa en Campus Cencosud'] = df_m['Nombre Capa en Campus Cencosud'].astype(str).str.strip()
        df_m['Nombre para el Reporte'] = df_m['Nombre para el Reporte'].astype(str).str.strip()

        # Normalizar nombres de columnas de la matriz
        map_u = df_m.set_index('Nombre Capa en UCENCO')['Nombre para el Reporte'].dropna().to_dict()
        map_c = df_m.set_index('Nombre Capa en Campus Cencosud')['Nombre para el Reporte'].dropna().to_dict()
        cols_reporte = df_m['Nombre para el Reporte'].unique().tolist()

    # 2. Unificar bases de capacitación
    bases = []
    if df_ucenco is not None and not df_ucenco.empty:
        df_u = df_ucenco.copy()
        df_u['Nombre_Estandar'] = df_u['Título de la capacitación'].map(map_u).fillna(df_u['Título de la capacitación'])
        bases.append(df_u)
        
    if df_campus is not None and not df_campus.empty:
        df_c = df_campus.copy()
        df_c['Nombre_Estandar'] = df_c['Título de la capacitación'].map(map_c).fillna(df_c['Título de la capacitación'])
        bases.append(df_c)

    if not bases:
        st.info("ℹ️ No hay datos de capacitación para consolidar.")
        return df_final
        
    df_all = pd.concat(bases, ignore_index=True)
    
    # 3. Aplicar Jerarquía
    # Definir Pesos
    pesos = {"Terminado": 3, "Desaprobado": 2, "Pendiente": 1, "No Aplica": 0}
    df_all['Peso'] = df_all['Estado_Calculado'].map(pesos).fillna(0)
    
    # Agrupar por ID y Nombre Estándar, tomar el máximo peso
    df_best = df_all.groupby(['ID_Clean', 'Nombre_Estandar'])['Peso'].max().reset_index()
    
    # Mapear de vuelta a texto
    inv_pesos = {v: k for k, v in pesos.items()}
    df_best['Estado_Final'] = df_best['Peso'].map(inv_pesos)
    
    # 4. Pivotar
    df_pivot = df_best.pivot(index='ID_Clean', columns='Nombre_Estandar', values='Estado_Final')
    df_pivot.reset_index(inplace=True)
    df_pivot.rename(columns={'ID_Clean': 'ID_Merge'}, inplace=True)
    
    # 5. Cruzar con el consolidado
    df_final['ID_Merge'] = df_final['ID Number'].apply(limpiar_id)
    df_final = pd.merge(df_final, df_pivot, on='ID_Merge', how='left')
    
    # 5.1 Refinar Lógica de "No Aplica" vs "Pendiente" usando Filtros Dinámicos
    # Identificar las columnas de cursos que se agregaron (son las que vienen de df_pivot)
    cursos_encontrados = [c for c in df_pivot.columns if c != 'ID_Merge']
    
    # Usar todas las columnas que deben estar en el reporte
    cursos_a_procesar = list(set(cursos_encontrados + cols_reporte))
    
    for curso in cursos_a_procesar:
        if curso not in df_final.columns:
            df_final[curso] = np.nan
            
        # 1. Rellenar con "Pendiente" (Status base para todos)
        df_final[curso] = df_final[curso].fillna("Pendiente")
        
        # 2. VALIDACIÓN DE EXCEPCIÓN (FILTROS DINÁMICOS) - PRIORIDAD MÁXIMA
        if not df_filtros.empty:
            mask_na = aplicar_filtros_dinamicos(df_final, df_filtros, modo="No Aplica", curso=curso)
            df_final.loc[mask_na, curso] = "No Aplica"
    st.write("📊 Calculando indicadores de cumplimiento...")
    
    # Identificar las columnas de cursos que se agregaron (son las que vienen de df_pivot)
    def calcular_resumen(row):
        # Valores de los cursos para esta fila
        vals = row[cursos_a_procesar]
        
        # Cursos aplicados = todos excepto "No Aplica"
        aplicados = (vals != "No Aplica").sum()
        
        # Cursos Pendientes = "Pendiente" o "Desaprobado"
        pendientes = vals.isin(["Pendiente", "Desaprobado"]).sum()
        
        # % Cumplimiento = ((Aplicados - Pendientes) / Aplicados) * 100
        if aplicados == 0:
            porcentaje = 0.0
        else:
            terminados = aplicados - pendientes
            porcentaje = (terminados / aplicados) * 100
            
        # ESTADO FINAL
        # - 100%: COMPLETADO
        # - 0% < x < 100%: INCOMPLETO
        # - 0%: SIN CURSO PERTINENTE (incluye aplicados = 0)
        if porcentaje == 100 and aplicados > 0:
            estado_f = "COMPLETADO"
        elif 0 <= porcentaje < 100:
            estado_f = "INCOMPLETO"
        else:
            estado_f = "SIN CURSO"
            
        return pd.Series([aplicados, pendientes, round(porcentaje, 1), estado_f])

    # Aplicar cálculo por fila
    if cursos_encontrados:
        df_resumen = df_final.apply(calcular_resumen, axis=1)
        df_final[['Cursos aplicados', 'Cursos Pendientes', '% Cumplimiento', 'ESTADO FINAL']] = df_resumen
    else:
        # Si no hay cursos, poner valores por defecto
        df_final['Cursos aplicados'] = 0
        df_final['Cursos Pendientes'] = 0
        df_final['% Cumplimiento'] = 0.0
        df_final['ESTADO FINAL'] = "SIN CURSO PERTINENTE"

    # Reordenar: Asegurar que las columnas de capacitación + resúmenes sigan el orden
    columnas_resumen = ['Cursos aplicados', 'Cursos Pendientes', '% Cumplimiento', 'ESTADO FINAL']
    if cols_reporte:
        # Solo reordenar las que realmente existen en el DataFrame
        orden_capacitacion = [c for c in cols_reporte if c in df_final.columns]
        
        # --- NUEVO: Ordenar por Índice Original de la Matriz ---
        if not df_m.empty:
            # Crear mapa de orden: {Curso: IndiceOriginal}
            # Usamos unique() para mantener el orden de aparición en la hoja de Excel
            nombres_ordenados = df_m['Nombre para el Reporte'].unique().tolist()
            map_orden_original = {curso: i for i, curso in enumerate(nombres_ordenados)}
            
            # Ordenar la lista de capacitación basándose en el Índice de la Matriz
            orden_capacitacion.sort(key=lambda x: map_orden_original.get(x, 9999))
        
        other_cols = [c for c in df_final.columns if c not in orden_capacitacion and c not in columnas_resumen]
        df_final = df_final[other_cols + orden_capacitacion + columnas_resumen]
        
    # Eliminar columna temporal
    if 'ID_Merge' in df_final.columns:
        df_final.drop(columns=['ID_Merge'], inplace=True)
        
    st.success(f"✅ Capacitaciones consolidadas con jerarquía e indicadores (Cursos: {len(cursos_encontrados)})")
    return df_final

def crear_hoja_resumen(wb, df_final, info_adicional):
    """
    Crea la hoja '% Cumplimiento' con el dashboard de tablas por Grupo TD.
    """
    try:
        df_m = info_adicional.get('df_matriz_cursos', pd.DataFrame())
        map_td = info_adicional.get('map_grupo_td', {})
        map_colores = info_adicional.get('map_colores_grupos', {})
        col_grupo = info_adicional.get('col_grupo_real', 'Grupo')

        if df_m.empty: return

        # Crear hoja al inicio
        ws = wb.create_sheet("% Cumplimiento", 0)
        
        # Estilos Base Dashboard
        dark_blue_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        light_blue_fill = PatternFill(start_color="3B5E91", end_color="3B5E91", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin', color="FFFFFF"),
                            right=Side(style='thin', color="FFFFFF"),
                            top=Side(style='thin', color="FFFFFF"),
                            bottom=Side(style='thin', color="FFFFFF"))

        # --- CABECERA BRANDING ---
        dir_assets = r"C:\Users\USUARIO\.gemini\antigravity\brain\f2ca4f0a-bd47-48d5-8d81-3b28911f9f0b"
        logos = {
            "cencosud": f"{dir_assets}\\cencosud_logo_placeholder_1774816848958.png",
            "wong": f"{dir_assets}\\wong_logo_placeholder_1774816871689.png",
            "metro": f"{dir_assets}\\metro_logo_placeholder_1774816892677.png"
        }
        
        try:
            if os.path.exists(logos["cencosud"]):
                img_c = Image(logos["cencosud"])
                img_c.width, img_c.height = 110, 50
                ws.add_image(img_c, "A2")
            if os.path.exists(logos["wong"]):
                img_w = Image(logos["wong"])
                img_w.width, img_w.height = 50, 50
                ws.add_image(img_w, "J2")
            if os.path.exists(logos["metro"]):
                img_m = Image(logos["metro"])
                img_m.width, img_m.height = 50, 50
                ws.add_image(img_m, "K2")
        except Exception as e: print(f"Logo err: {e}")

        ws.merge_cells("C2:I4")
        cell_title = ws["C2"]
        cell_title.value = "REPORTE DE CURSOS NORMATIVOS 2026"
        cell_title.fill = PatternFill(start_color="1F618D", end_color="1F618D", fill_type="solid")
        cell_title.font = Font(color="FFFFFF", bold=True, size=22)
        cell_title.alignment = center_align
        
        # Subtítulo (Banner Verde)
        ws.merge_cells("A6:K6")
        ws["A6"] = f"% Cumplimiento por curso normativo {datetime.now().year}"
        ws["A6"].fill = PatternFill(start_color="248232", end_color="248232", fill_type="solid")
        ws["A6"].font = Font(color="FFFFFF", bold=True, size=14)
        ws["A6"].alignment = Alignment(horizontal="left", vertical="center")
        
        # 1. TABLAS POR GRUPO TD (RESTAURADAS A LA POSICIÓN ORIGINAL)
        curr_row = 8
        unique_tds = sorted(list(set(map_td.values())), key=lambda x: str(x))
        for td in unique_tds:
            nombres_ordenados = df_m['Nombre para el Reporte'].unique()
            cursos_td = [c for c in nombres_ordenados if map_td.get(c) == td]
            if not cursos_td: continue
            cursos_reales = [c for c in cursos_td if c in df_final.columns]
            if not cursos_reales: continue
            
            # Fila de Grupos
            col_idx = 2
            while (col_idx - 2) < len(cursos_reales):
                curso = cursos_reales[col_idx - 2]
                grp_name = map_c_g = df_m[df_m['Nombre para el Reporte'] == curso][col_grupo].fillna("GENERAL").iloc[0]
                bg_color = map_colores.get(str(grp_name).upper(), "002060")
                start_col = col_idx
                while (col_idx - 2) < len(cursos_reales) and df_m[df_m['Nombre para el Reporte'] == cursos_reales[col_idx-2]][col_grupo].fillna("GENERAL").iloc[0] == grp_name:
                    cell = ws.cell(row=curr_row, column=col_idx)
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    cell.font = white_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    col_idx += 1
                if col_idx - 1 > start_col:
                    ws.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=col_idx-1)
                ws.cell(row=curr_row, column=start_col).value = grp_name
            
            curr_row += 1
            ws.cell(row=curr_row, column=1).value = "Estado"
            ws.cell(row=curr_row, column=1).fill = dark_blue_fill
            ws.cell(row=curr_row, column=1).font = white_font
            ws.cell(row=curr_row, column=1).border = thin_border
            
            col_idx = 2
            for curso in cursos_reales:
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.value = curso
                grp_name = df_m[df_m['Nombre para el Reporte'] == curso][col_grupo].fillna("GENERAL").iloc[0]
                bg_color = map_colores.get(str(grp_name).upper(), "002060")
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                cell.font = Font(color="FFFFFF", size=9, bold=True)
                cell.alignment = center_align
                cell.border = thin_border
                col_idx += 1
                
            for label in ["Pendiente", "Desaprobado", "Terminado", "Total", "% Avance"]:
                curr_row += 1
                cell_label = ws.cell(row=curr_row, column=1)
                cell_label.value = label
                cell_label.fill = dark_blue_fill if label != "Total" else light_blue_fill
                cell_label.font = white_font
                cell_label.border = thin_border
                col_idx = 2
                for curso in cursos_reales:
                    cell_val = ws.cell(row=curr_row, column=col_idx)
                    counts = df_final[curso].value_counts()
                    if label == "Pendiente": val = counts.get("Pendiente", 0) + counts.get("INCOMPLETO", 0)
                    elif label == "Desaprobado": val = counts.get("Desaprobado", 0)
                    elif label == "Terminado": val = counts.get("Terminado", 0) + counts.get("COMPLETADO", 0)
                    elif label == "Total": 
                        # Total real = Total - No Aplica (Solo poblacion asignada)
                        val = counts.sum() - counts.get("No Aplica", 0)
                    elif label == "% Avance":
                        d = counts.get("Terminado", 0) + counts.get("COMPLETADO", 0)
                        # El total para el % tambien debe excluir No Aplica
                        t = counts.sum() - counts.get("No Aplica", 0)
                        val = (d/t) if t > 0 else 0
                        cell_val.number_format = '0.0%'
                    cell_val.value = val
                    cell_val.alignment = Alignment(horizontal="center")
                    cell_val.border = Border(left=Side(style='thin', color="D9D9D9"), right=Side(style='thin', color="D9D9D9"), top=Side(style='thin', color="D9D9D9"), bottom=Side(style='thin', color="D9D9D9"))
                    col_idx += 1
            curr_row += 3

        # 2. TABLA RESUMEN JGHZ / JGH (TIPO DINÁMICA - MEJORADA)
        curr_row += 2
        header_p = ws.cell(row=curr_row, column=1)
        header_p.value = "NORMATIVOS SEGÚN JGHZ POR CUMPLIMIENTO DE CURSOS"
        header_p.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_p.font = Font(color="FFFFFF", bold=True, size=12)
        header_p.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{curr_row}:E{curr_row}")
        
        # --- PREPARACIÓN DE DATOS (PARA PIVOT NATIVO) ---
        df_para_resumen = df_final[df_final['ESTADO FINAL'].isin(['COMPLETADO', 'INCOMPLETO'])].copy()

        # --- CREAR PIVOT NATIVO EN EL DASHBOARD ---

        # --- CREAR PIVOT NATIVO EN EL DASHBOARD (Usando el reporte principal como origen) ---
        from openpyxl.pivot.table import PivotTable, PivotField, RowField, ColumnField, DataField
        from openpyxl.worksheet.table import TableStyleInfo
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import Reference
        
        try:
            # 1. Identificar la hoja de origen y rango
            ws_source = wb[wb.sheetnames[-1]]
            max_r = ws_source.max_row
            num_c = ws_source.max_column
            
            # Buscar índices de columnas JGHZ, JGH, ESTADO FINAL en la fila 2
            idx_jghz = idx_jgh = idx_ef = None
            for c in range(1, num_c + 1):
                val = str(ws_source.cell(row=2, column=c).value).strip().upper()
                if "JGH ZONAL" in val: idx_jghz = c
                elif "JGH" == val: idx_jgh = c
                elif "ESTADO FINAL" in val: idx_ef = c
            
            if idx_jghz and idx_jgh and idx_ef:
                # Rango del Pivot (Header en fila 2)
                pivot_source = Reference(ws_source, min_row=2, min_col=1, max_row=max_r, max_col=num_c)
                
                pivot = PivotTable(pivotFieldName="ResumenJefaturas", cacheDefinition=None)
                pivot.name = "PivotJefaturas"
                pivot.dataCaption = "Resumen Jefaturas"
                
                # Índices 0-based para el pivot (Basado en el rango que empieza en col 1)
                # 1. JGHZ como Fila 0
                pivot.pivotFields.append(PivotField(axis="axisRow", showAllItems=True))
                pivot.rowFields.append(RowField(indx=idx_jghz-1))
                
                # 2. JGH como Fila 1
                pivot.pivotFields.append(PivotField(axis="axisRow", showAllItems=True))
                pivot.rowFields.append(RowField(indx=idx_jgh-1))
                
                # 3. ESTADO FINAL como Columna y Valor
                pivot.pivotFields.append(PivotField(axis="axisCol", showAllItems=True))
                pivot.colFields.append(ColumnField(indx=idx_ef-1))
                
                # Valor (Cuenta de ESTADO FINAL)
                measure = DataField(name="Cuenta de ESTADO FINAL", fld=idx_ef-1, subtotal="count")
                pivot.dataFields.append(measure)
                
                # Estilo
                pivot.tableStyleInfo = TableStyleInfo(name="PivotStyleLight16")
                
                # Añadir al Dashboard (un poco debajo de las tablas fijas)
                ws.add_pivot_table(pivot, f"A{curr_row+5}", pivot_source)
        except Exception as e:
            print(f"Error creando pivot nativo: {e}")

        # Ancho de columnas
        ws.column_dimensions['A'].width = 15
        for i in range(2, 50):
            ws.column_dimensions[get_column_letter(i)].width = 15

    except Exception as e:
        print(f"Error creando dashboard: {e}")

def aplicar_estilos_custom(file_path, info_adicional, df_final=None):
    """
    Aplica el diseño corporativo y crea el dashboard de cumplimiento.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        
        # 0. Crear Dashboard si hay datos
        if df_final is not None:
            crear_hoja_resumen(wb, df_final, info_adicional)
        
        # Seleccionar hoja de reporte (asumimos la última cargada inicialmente)
        ws_name = next((s for s in wb.sheetnames if 'reporte' in s.lower()), wb.sheetnames[-1])
        ws = wb[ws_name]
        
        # 1. Obtener Matriz de Grupos
        df_m = info_adicional.get('df_matriz_cursos', pd.DataFrame())
        col_grupo = info_adicional.get('col_grupo_real', 'Grupo')
        
        if df_m.empty:
            map_grupo = {}
        else:
            # Limpiar y mapear
            df_m['Nombre para el Reporte'] = df_m['Nombre para el Reporte'].astype(str).str.strip()
            if col_grupo in df_m.columns:
                df_m[col_grupo] = df_m[col_grupo].astype(str).str.strip().str.upper()
                map_grupo = df_m.set_index('Nombre para el Reporte')[col_grupo].dropna().to_dict()
            else:
                map_grupo = {}
            
        # 2. Preparar nueva fila para Grupos
        ws.insert_rows(1)
        
        # Mapeo de Colores (Dinámico desde la matriz)
        color_map = info_adicional.get('map_colores_grupos', {})
        navy_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        
        # Estilos Base
        white_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        white_border = Border(left=Side(style='thin', color="FFFFFF"), 
                             right=Side(style='thin', color="FFFFFF"), 
                             top=Side(style='thin', color="FFFFFF"), 
                             bottom=Side(style='thin', color="FFFFFF"))
        
        # 3. Lógica de Grupos y Estilo de Fila 1
        num_cols = ws.max_column
        col_idx = 1
        while col_idx <= num_cols:
            h_text = ws.cell(row=2, column=col_idx).value
            raw_group = str(map_grupo.get(h_text, "")).strip().upper()
            
            if raw_group and raw_group != "NAN" and raw_group != "SIN GRUPO":
                # Determinar color del grupo
                hex_color = color_map.get(raw_group, "002060") # Default navy
                group_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                
                # Fila 1 Cabecera
                ws.cell(row=1, column=col_idx).fill = group_fill
                ws.cell(row=1, column=col_idx).font = white_font
                ws.cell(row=1, column=col_idx).alignment = center_align
                ws.cell(row=1, column=col_idx).border = white_border

                start_col = col_idx
                # Avanzar mientras el grupo sea el mismo
                while col_idx < num_cols and str(map_grupo.get(ws.cell(row=2, column=col_idx + 1).value)).strip().upper() == raw_group:
                    col_idx += 1
                    ws.cell(row=1, column=col_idx).fill = group_fill
                    ws.cell(row=1, column=col_idx).font = white_font
                    ws.cell(row=1, column=col_idx).alignment = center_align
                    ws.cell(row=1, column=col_idx).border = white_border
                
                # Combinar y Agrupar (Outline)
                if col_idx - start_col >= 0:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col_idx)
                    ws.cell(row=1, column=start_col).value = raw_group
                    ws.column_dimensions.group(get_column_letter(start_col), get_column_letter(col_idx), outline_level=1, hidden=False)
                
                col_idx += 1
            else:
                col_idx += 1

        # 4. Estilo de Cabeceras (Fila 2) y Filtros
        for c in range(1, num_cols + 1):
            cell_f2 = ws.cell(row=2, column=c)
            # Determinar color de fondo para la fila 2 (un poco más oscuro o igual)
            h_text_f2 = cell_f2.value
            raw_group_f2 = str(map_grupo.get(h_text_f2, "")).strip().upper()
            hex_color_f2 = color_map.get(raw_group_f2, "002060")
            
            cell_f2.fill = PatternFill(start_color=hex_color_f2, end_color=hex_color_f2, fill_type="solid")
            cell_f2.font = white_font
            cell_f2.alignment = center_align
            cell_f2.border = white_border

        # Activar Auto-Filtros en la fila 2
        ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}2"

        # 5. Formato Condicional (Celdas de Estado)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gray_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Reglas basadas en el texto exacto
        rules = [
            ("COMPLETADO", green_fill),
            ("INCOMPLETO", red_fill),
            ("No Aplica", gray_fill),
            ("SIN CURSO", gray_fill)
        ]
        
        for text, fill in rules:
            ws.conditional_formatting.add(
                f"A3:{get_column_letter(num_cols)}{ws.max_row}",
                CellIsRule(operator='equal', formula=[f'"{text}"'], fill=fill)
            )

        # 6. Auto-ajuste de columnas y congelar panes
        for col in ws.columns:
            max_length = 0
            # Usar get_column_letter(col[0].column) es más seguro con celdas combinadas
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 45)

        ws.freeze_panes = "A3"
        
        wb.save(file_path)
        return True
    except Exception as e:
        st.error(f"⚠️ Error aplicando estilos al Excel: {e}")
        return False
